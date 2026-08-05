"""Group booking — the passenger list arrives as a spreadsheet.

A party of eighty is not typed into a form. The merchant downloads a template,
fills one row per traveller, and uploads it; those rows become the booking's
passengers. **Nothing after the import differs** — same Manager approval, same
issuance, same wallet rules, same statuses. This module owns exactly three
things: producing the template, reading a filled one back, and judging it.

WHY THE SHEET IS FLAT
Journey and booking columns repeat on every row even though they describe the
booking rather than the traveller. That is deliberate and it is what merchants
already do in their own spreadsheets: one row is one person, and a row that
carries its own context can be sorted, filtered and pasted without losing it.
The cost is that the same journey is stated N times and the rows can disagree —
:func:`_read_journey` is where that is caught rather than silently resolved.

WHY THE FORM STILL OWNS THE ITINERARY
The booking is raised from the journey fields on screen, not from the sheet.
The sheet's journey columns are validated and recorded (``imported_journey``)
so a disagreement is *visible*, but they never overwrite what the merchant
confirmed on the form. Letting an uploaded file rewrite the itinerary would mean
a booking could travel somewhere nobody typed.

VALIDATION IS PER ROW, AND IT DOES NOT STOP AT THE FIRST FAILURE
A merchant fixing a 200-row sheet one error per upload would upload forty times.
Every row is judged, every problem is collected with its **Excel row number**,
and the whole list comes back at once. ``partial`` is a real outcome: the sheet
parsed, some rows are unusable, and the merchant decides whether to fix them or
replace the file. Only a wholly ``valid`` import may become a booking.

TRUST
The uploaded bytes are attacker-controlled. The declared filename and content
type are used for display and nothing else; the format is decided by sniffing
the magic bytes (:func:`_sniff_format`), the size is capped before parsing, and
a workbook is opened with formulas discarded (``data_only``) so a sheet cannot
smuggle a computation into a cell this module then reads as a name.
"""
from __future__ import annotations

import datetime
import hashlib
import io
import re
import tempfile
import uuid
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile, status as http_status
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models_v2 import (
    Gender,
    GroupBookingImport,
    GroupImportStatus,
    PassengerType,
    ServiceRequest,
    User,
)
from app.services import storage

# ---------------------------------------------------------------------------
# What the template contains
# ---------------------------------------------------------------------------
#: Journey columns, on every row. Named exactly as the merchant sees them —
#: these strings are the contract between the file they filled and the parser,
#: so renaming one is a breaking change to every template already downloaded.
JOURNEY_COLUMNS: tuple[str, ...] = (
    "Origin City",
    "Destination City",
    "Airline",
    "Airline Number",
    "Travel Date",
    "Preferred Hour",
    "Preferred Minute",
    "AM/PM",
)

#: Added for a Round Trip Group only. A One Way Group template does not carry
#: them at all: an empty return column invites a return date on a one-way
#: booking, which the itinerary validator would then have to refuse.
RETURN_COLUMNS: tuple[str, ...] = (
    "Return Date",
    "Return Preferred Hour",
    "Return Preferred Minute",
    "Return AM/PM",
)

PASSENGER_COLUMNS: tuple[str, ...] = (
    "Passenger Name",
    "Gender",
    "Date of Birth",
    "Nationality",
    "Passport Number",
    "Passport Expiry",
)

BOOKING_COLUMNS: tuple[str, ...] = (
    "Booking Class",
    "Adults",
    "Children",
    "Infants",
    "Client Fare",
    "Notes",
)

#: Columns a row cannot be imported without. Everything else is optional —
#: passports are only mandatory on an international sector, and that is decided
#: by the booking form (which knows the countries), not by this sheet.
REQUIRED_COLUMNS: frozenset[str] = frozenset(
    {
        "Origin City",
        "Destination City",
        "Airline",
        "Airline Number",
        "Travel Date",
        "Passenger Name",
    }
)

ONE_WAY_GROUP = "one_way_group"
ROUND_TRIP_GROUP = "round_trip_group"
JOURNEY_TYPES = (ONE_WAY_GROUP, ROUND_TRIP_GROUP)

#: Mirrors the merchant form's dropdown (CR-5). Offered as an Excel dropdown so
#: the desk does not receive "buisness" forty times.
TRAVEL_CLASSES = ("Economy", "Premium Economy", "Business", "First Class")
GENDERS = ("Male", "Female", "Other")
MERIDIEMS = ("AM", "PM")

#: 10 MB, matching what the upload card advertises. Checked while streaming, so
#: an oversized file never lands in memory whole.
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
_CHUNK = 64 * 1024

#: What the stored manifest is served back as. The bytes kept are the
#: merchant's original file, unmodified — this is the type the download
#: endpoint declares, and a legacy .xls opens from it perfectly well in Excel.
_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#: openpyxl reads the zip-based format; xlrd 2.x reads only the OLE2 one. The
#: signatures are how a file is routed, because the extension is a claim.
_XLSX_MAGIC = b"PK\x03\x04"
_XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

#: A flight number is 2-3 letters (IATA/ICAO carrier) then 1-4 digits, with an
#: optional trailing operational suffix. Deliberately permissive about spacing.
_FLIGHT_RE = re.compile(r"^[A-Z0-9]{2,3}\s*\d{1,4}[A-Z]?$", re.IGNORECASE)

#: Cap on rows read. A group is a party on an aircraft, not a database dump, and
#: an unbounded loop over a hostile sheet is the easy way to spend all the RAM.
MAX_ROWS = 1000


def columns_for(journey_type: str) -> tuple[str, ...]:
    """Every column of the template, in order, for one journey type."""
    _assert_journey_type(journey_type)
    cols = list(JOURNEY_COLUMNS)
    if journey_type == ROUND_TRIP_GROUP:
        cols += list(RETURN_COLUMNS)
    cols += list(PASSENGER_COLUMNS)
    cols += list(BOOKING_COLUMNS)
    return tuple(cols)


def _assert_journey_type(journey_type: str) -> None:
    if journey_type not in JOURNEY_TYPES:
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail=f"Unknown group journey type: {journey_type!r}",
        )


# ---------------------------------------------------------------------------
# Building the template
# ---------------------------------------------------------------------------
def build_template(journey_type: str) -> bytes:
    """Produce the .xlsx the merchant fills in.

    Everything that can be constrained in the file itself is: dates are real
    date cells with a visible format, the three enumerated columns are
    dropdowns, and the counts are whole numbers. A constraint expressed in the
    file is one the merchant cannot get wrong before they upload, which is worth
    far more than the same rule stated in an error message afterwards.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    _assert_journey_type(journey_type)
    cols = columns_for(journey_type)

    wb = Workbook()
    ws = wb.active
    ws.title = "Passengers"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    required_fill = PatternFill("solid", fgColor="0F2F4F")

    for idx, name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = header_font
        # Required columns get a darker header. The legend sheet explains it —
        # a merchant should be able to see what is mandatory without reading it.
        cell.fill = required_fill if name in REQUIRED_COLUMNS else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = max(14, len(name) + 4)

    ws.freeze_panes = "A2"

    # One filled row, so the expected shape of every value is unambiguous.
    sample_date = datetime.date.today() + datetime.timedelta(days=30)
    sample: dict[str, Any] = {
        "Origin City": "Hyderabad",
        "Destination City": "Delhi",
        "Airline": "Air India",
        "Airline Number": "AI217",
        "Travel Date": sample_date,
        "Preferred Hour": 9,
        "Preferred Minute": 0,
        "AM/PM": "AM",
        "Return Date": sample_date + datetime.timedelta(days=7),
        "Return Preferred Hour": 6,
        "Return Preferred Minute": 30,
        "Return AM/PM": "PM",
        "Passenger Name": "John Doe",
        "Gender": "Male",
        "Date of Birth": datetime.date(1990, 4, 18),
        "Nationality": "Indian",
        "Passport Number": "P1234567",
        "Passport Expiry": datetime.date(sample_date.year + 5, 4, 18),
        "Booking Class": "Economy",
        "Adults": 1,
        "Children": 0,
        "Infants": 0,
        "Client Fare": 20000,
        "Notes": "Seats together if possible",
    }
    for idx, name in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=idx, value=sample.get(name))
        if isinstance(sample.get(name), datetime.date):
            cell.number_format = "DD/MM/YYYY"

    # Dropdowns. Applied over a generous row range so they survive the merchant
    # pasting several hundred rows in below the sample.
    last_row = MAX_ROWS + 1
    dropdowns = {
        "Gender": GENDERS,
        "AM/PM": MERIDIEMS,
        "Return AM/PM": MERIDIEMS,
        "Booking Class": TRAVEL_CLASSES,
    }
    for name, options in dropdowns.items():
        if name not in cols:
            continue
        letter = get_column_letter(cols.index(name) + 1)
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(options) + '"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.error = f"Choose one of: {', '.join(options)}"
        dv.errorTitle = name
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{last_row}")

    # Date columns: formatted, and constrained to something plausible so a
    # mis-keyed year is caught in Excel rather than three screens later.
    for name in ("Travel Date", "Return Date", "Date of Birth", "Passport Expiry"):
        if name not in cols:
            continue
        letter = get_column_letter(cols.index(name) + 1)
        for row in range(2, last_row + 1):
            ws[f"{letter}{row}"].number_format = "DD/MM/YYYY"
        dv = DataValidation(
            type="date",
            operator="between",
            formula1=datetime.date(1900, 1, 1),
            formula2=datetime.date(2100, 12, 31),
            allow_blank=True,
        )
        dv.error = "Enter a real date, formatted DD/MM/YYYY"
        dv.errorTitle = name
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{last_row}")

    for name in ("Preferred Hour", "Return Preferred Hour"):
        if name not in cols:
            continue
        letter = get_column_letter(cols.index(name) + 1)
        dv = DataValidation(
            type="whole", operator="between", formula1=1, formula2=12, allow_blank=True
        )
        dv.error = "The hour is 1-12; use the AM/PM column for the rest"
        dv.errorTitle = name
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{last_row}")

    for name in ("Preferred Minute", "Return Preferred Minute"):
        if name not in cols:
            continue
        letter = get_column_letter(cols.index(name) + 1)
        dv = DataValidation(
            type="whole", operator="between", formula1=0, formula2=59, allow_blank=True
        )
        dv.error = "Minutes are 0-59"
        dv.errorTitle = name
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{last_row}")

    for name in ("Adults", "Children", "Infants"):
        letter = get_column_letter(cols.index(name) + 1)
        dv = DataValidation(
            type="whole", operator="between", formula1=0, formula2=99, allow_blank=True
        )
        dv.error = "Enter a whole number of passengers"
        dv.errorTitle = name
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{last_row}")

    _add_instructions_sheet(wb, journey_type, cols)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_instructions_sheet(wb, journey_type: str, cols: tuple[str, ...]) -> None:
    """A second sheet explaining the first.

    The merchant filling this in is not the person who read the spec. Column
    meanings, the mandatory set and the date format live in the file itself so
    the answer travels with it.
    """
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 82

    label = "Round Trip Group" if journey_type == ROUND_TRIP_GROUP else "One Way Group"
    lines: list[tuple[str, str]] = [
        (f"{label} passenger list", ""),
        ("", ""),
        ("One row per traveller", "Journey and booking columns repeat on every row. Keep them identical down the sheet."),
        ("Do not rename columns", "The header row is how the file is read. Renaming, reordering or deleting a column will fail the upload."),
        ("Dates", "DD/MM/YYYY. Type a real date — the cells are date-formatted, so Excel will show it back to you correctly if it took it."),
        ("Time", "Preferred Hour is 1-12, Preferred Minute is 0-59, and AM/PM decides the rest."),
        ("Passenger Name", "Full name, as printed in the passport. First name and surname, separated by a space."),
        ("Mandatory columns", ", ".join(c for c in cols if c in REQUIRED_COLUMNS)),
        ("Passport", "Required only for international sectors. Leave blank for a domestic group."),
        ("Client Fare", "What you have quoted your own customer. Optional, and never used to bill you."),
        ("Row limit", f"Up to {MAX_ROWS} passengers per upload, and up to 10 MB."),
    ]
    for row, (head, body) in enumerate(lines, start=1):
        a = ws.cell(row=row, column=1, value=head)
        b = ws.cell(row=row, column=2, value=body)
        a.font = Font(bold=True, size=13 if row == 1 else 11)
        b.alignment = Alignment(wrap_text=True, vertical="top")


def template_filename(journey_type: str) -> str:
    kind = "round-trip" if journey_type == ROUND_TRIP_GROUP else "one-way"
    return f"jackpots-group-booking-{kind}-template.xlsx"


# ---------------------------------------------------------------------------
# Reading a filled template back
# ---------------------------------------------------------------------------
def _sniff_format(head: bytes) -> str:
    """Decide the format from the bytes, never from the filename."""
    if head.startswith(_XLSX_MAGIC):
        return "xlsx"
    if head.startswith(_XLS_MAGIC):
        return "xls"
    raise HTTPException(
        status_code=http_status.HTTP_400_BAD_REQUEST,
        detail=(
            "That is not an Excel workbook. Upload the .xlsx template you "
            "downloaded — a CSV or a renamed file will not do."
        ),
    )


def _rows_from_xlsx(data: bytes) -> list[list[Any]]:
    from openpyxl import load_workbook

    try:
        # data_only: a cell holding a formula yields its cached value, never the
        # formula text. read_only keeps a large sheet off the heap.
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises a zoo of exception types
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail=f"That workbook could not be opened: {exc}",
        ) from exc

    ws = wb["Passengers"] if "Passengers" in wb.sheetnames else wb.worksheets[0]
    rows = [list(r) for r in ws.iter_rows(max_row=MAX_ROWS + 1, values_only=True)]
    wb.close()
    return rows


def _rows_from_xls(data: bytes) -> list[list[Any]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - dependency is in requirements
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail=(
                "Legacy .xls files cannot be read on this server. Open the file "
                "and use Save As -> Excel Workbook (.xlsx)."
            ),
        ) from exc

    try:
        book = xlrd.open_workbook(file_contents=data)
    except Exception as exc:
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail=f"That workbook could not be opened: {exc}",
        ) from exc

    names = book.sheet_names()
    sheet = book.sheet_by_name("Passengers") if "Passengers" in names else book.sheet_by_index(0)

    rows: list[list[Any]] = []
    for r in range(min(sheet.nrows, MAX_ROWS + 1)):
        row: list[Any] = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            value = cell.value
            # xlrd hands back dates as floats in the workbook's epoch; convert
            # here so the rest of this module sees one date type.
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    y, mo, d, h, mi, s = xlrd.xldate_as_tuple(value, book.datemode)
                    value = datetime.date(y, mo, d) if y else None
                except Exception:
                    value = None
            row.append(value)
        rows.append(row)
    return rows


def parse_rows(data: bytes) -> list[list[Any]]:
    """Every row of the workbook, header included, as plain Python values."""
    fmt = _sniff_format(data[:8])
    return _rows_from_xlsx(data) if fmt == "xlsx" else _rows_from_xls(data)


# ---------------------------------------------------------------------------
# Coercion helpers — a spreadsheet cell is whatever the merchant typed
# ---------------------------------------------------------------------------
def _text(value: Any) -> str:
    """A trimmed string, with Excel's numeric surprises undone.

    A passport number typed into a General cell comes back as ``1234567.0``.
    Rendering that straight to text stores a passport that will not match the
    document, so an integral float is narrowed before it is stringified.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    return str(value).strip()


def _int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        if isinstance(value, str):
            value = value.strip()
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        # Strip thousands separators and a currency symbol — merchants paste
        # from their own quoting sheets, which are formatted for humans.
        cleaned = re.sub(r"[^\d.\-]", "", str(value))
        return Decimal(cleaned) if cleaned not in ("", "-", ".") else None
    except (InvalidOperation, ValueError):
        return None


#: Accepted date spellings, in the order they are tried. Day-first throughout —
#: the template says DD/MM/YYYY and the merchants are Indian, so 08/04/2026 is
#: the 8th of April. ISO is accepted because it is unambiguous.
_DATE_FORMATS = ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d %b %Y", "%d %B %Y")


def _date(value: Any) -> datetime.date | None:
    """A real date, or None. Excel date cells arrive already typed."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    # A bare serial number, from a cell formatted as General.
    if isinstance(value, (int, float)):
        try:
            return (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(value)))
        except (OverflowError, ValueError):
            return None
    text = str(value).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _split_name(full: str) -> tuple[str, str] | None:
    """"John Doe" -> ("John", "Doe"). None when there is only one part.

    Split on the LAST space, so "Maria del Carmen Rodriguez" keeps its given
    names together and takes "Rodriguez" as the surname — which is what an
    airline manifest wants and what a passport's machine-readable zone says.
    """
    parts = [p for p in re.split(r"\s+", full.strip()) if p]
    if len(parts) < 2:
        return None
    return " ".join(parts[:-1]), parts[-1]


def _gender(value: str) -> Gender | None:
    cleaned = value.strip().casefold()
    if cleaned in ("m", "male"):
        return Gender.MALE
    if cleaned in ("f", "female"):
        return Gender.FEMALE
    if cleaned in ("o", "other", "x"):
        return Gender.OTHER
    return None


def _time_24h(hour: Any, minute: Any, meridiem: Any) -> str | None:
    """The template's 12-hour triple as the ``HH:MM`` the API takes."""
    h = _int(hour)
    m = _int(minute) or 0
    mer = _text(meridiem).strip().upper()
    if h is None:
        return None
    if not (1 <= h <= 12) or not (0 <= m <= 59) or mer not in MERIDIEMS:
        return None
    if mer == "AM":
        h = 0 if h == 12 else h
    else:
        h = 12 if h == 12 else h + 12
    return f"{h:02d}:{m:02d}"


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
class _Err(dict):
    """One row-level problem. A dict so it lands in JSONB unchanged."""

    def __init__(self, row: int, message: str, column: str | None = None):
        super().__init__(row=row, column=column, message=message)


def validate_sheet(rows: list[list[Any]], journey_type: str) -> dict[str, Any]:
    """Judge a parsed workbook.

    Returns the journey the sheet declared, the passengers that survived, and
    every problem found — all of them, with the merchant's own row numbers.
    """
    _assert_journey_type(journey_type)
    expected = columns_for(journey_type)
    errors: list[_Err] = []

    if not rows:
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail="That workbook is empty.",
        )

    # --- the header -------------------------------------------------------
    header = [_text(c) for c in rows[0]]
    index = {name: i for i, name in enumerate(header) if name}
    missing = [c for c in expected if c not in index]
    if missing:
        # A missing column is not a row-level problem — nothing can be read at
        # all — so this is the one validation that refuses the whole file.
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail=(
                "This sheet is missing required columns: "
                + ", ".join(missing)
                + ". Download the template again and paste your rows into it."
            ),
        )

    def cell(row: list[Any], column: str) -> Any:
        i = index.get(column)
        return row[i] if i is not None and i < len(row) else None

    # --- the rows ---------------------------------------------------------
    passengers: list[dict[str, Any]] = []
    journey_votes: list[dict[str, Any]] = []
    seen_passports: dict[str, int] = {}
    today = datetime.date.today()
    data_rows = 0

    for offset, row in enumerate(rows[1:], start=0):
        excel_row = offset + 2  # 1-based, and the header is row 1
        if all(_text(c) == "" for c in row):
            continue  # a blank spacer row is not an error
        data_rows += 1
        row_errors: list[_Err] = []

        # Required values
        for column in expected:
            if column in REQUIRED_COLUMNS and _text(cell(row, column)) == "":
                row_errors.append(_Err(excel_row, f"{column} is missing.", column))

        origin = _text(cell(row, "Origin City"))
        destination = _text(cell(row, "Destination City"))
        if origin and destination and origin.casefold() == destination.casefold():
            row_errors.append(
                _Err(excel_row, "Origin and Destination cannot be the same.")
            )

        airline = _text(cell(row, "Airline"))
        flight = _text(cell(row, "Airline Number"))
        if flight and not _FLIGHT_RE.match(flight.replace(" ", "")):
            row_errors.append(
                _Err(
                    excel_row,
                    f"{flight!r} is not a valid airline number — expected something "
                    "like AI217 or 6E456.",
                    "Airline Number",
                )
            )

        travel_date = _date(cell(row, "Travel Date"))
        if _text(cell(row, "Travel Date")) and travel_date is None:
            row_errors.append(_Err(excel_row, "Travel Date is invalid.", "Travel Date"))
        elif travel_date and travel_date < today:
            row_errors.append(
                _Err(excel_row, "Travel Date is in the past.", "Travel Date")
            )

        return_date = None
        if journey_type == ROUND_TRIP_GROUP:
            return_date = _date(cell(row, "Return Date"))
            if return_date is None:
                row_errors.append(
                    _Err(excel_row, "Return Date is missing or invalid.", "Return Date")
                )
            elif travel_date and return_date <= travel_date:
                row_errors.append(
                    _Err(
                        excel_row,
                        "Return Date must be after the Travel Date.",
                        "Return Date",
                    )
                )

        # Passenger
        name = _text(cell(row, "Passenger Name"))
        split = _split_name(name) if name else None
        if name and split is None:
            row_errors.append(
                _Err(
                    excel_row,
                    "Passenger Name needs a first name and a surname.",
                    "Passenger Name",
                )
            )

        gender_raw = _text(cell(row, "Gender"))
        gender = _gender(gender_raw) if gender_raw else None
        if gender_raw and gender is None:
            row_errors.append(
                _Err(excel_row, f"{gender_raw!r} is not a valid Gender.", "Gender")
            )

        dob = _date(cell(row, "Date of Birth"))
        if _text(cell(row, "Date of Birth")) and dob is None:
            row_errors.append(
                _Err(excel_row, "Date of Birth is invalid.", "Date of Birth")
            )
        elif dob and dob > today:
            row_errors.append(
                _Err(excel_row, "Date of Birth is in the future.", "Date of Birth")
            )

        passport = _text(cell(row, "Passport Number"))
        if passport:
            # Duplicates WITHIN this sheet. Two rows sharing a passport is one
            # traveller entered twice — the airline would reject the second at
            # ticketing, and it is far cheaper to say so now.
            previous = seen_passports.get(passport.casefold())
            if previous:
                row_errors.append(
                    _Err(
                        excel_row,
                        f"Passport Number {passport} is already used on row {previous}.",
                        "Passport Number",
                    )
                )
            else:
                seen_passports[passport.casefold()] = excel_row

        passport_expiry = _date(cell(row, "Passport Expiry"))
        if _text(cell(row, "Passport Expiry")) and passport_expiry is None:
            row_errors.append(
                _Err(excel_row, "Passport Expiry is invalid.", "Passport Expiry")
            )
        elif passport_expiry and travel_date and passport_expiry < travel_date:
            row_errors.append(
                _Err(
                    excel_row,
                    "Passport Expiry is before the Travel Date.",
                    "Passport Expiry",
                )
            )

        # Counts. Present on every row but describing the booking, so they are
        # validated for sanity here and reconciled across the sheet below.
        adults = _int(cell(row, "Adults"))
        children = _int(cell(row, "Children"))
        infants = _int(cell(row, "Infants"))
        for label, value in (("Adults", adults), ("Children", children), ("Infants", infants)):
            if _text(cell(row, label)) and value is None:
                row_errors.append(
                    _Err(excel_row, f"{label} must be a whole number.", label)
                )
            elif value is not None and value < 0:
                row_errors.append(_Err(excel_row, f"{label} cannot be negative.", label))

        if row_errors:
            errors.extend(row_errors)
            continue

        journey_votes.append(
            {
                "_row": excel_row,
                "origin_city": origin,
                "destination_city": destination,
                "airline": airline,
                "flight_number": flight.upper(),
                "travel_date": travel_date.isoformat() if travel_date else None,
                "preferred_time": _time_24h(
                    cell(row, "Preferred Hour"),
                    cell(row, "Preferred Minute"),
                    cell(row, "AM/PM"),
                ),
                "return_date": return_date.isoformat() if return_date else None,
                "return_preferred_time": (
                    _time_24h(
                        cell(row, "Return Preferred Hour"),
                        cell(row, "Return Preferred Minute"),
                        cell(row, "Return AM/PM"),
                    )
                    if journey_type == ROUND_TRIP_GROUP
                    else None
                ),
                "travel_class": _text(cell(row, "Booking Class")) or "Economy",
                "adults": adults,
                "children": children,
                "infants": infants,
                "client_fare": (
                    str(_decimal(cell(row, "Client Fare")))
                    if _decimal(cell(row, "Client Fare")) is not None
                    else None
                ),
                "notes": _text(cell(row, "Notes")) or None,
            }
        )

        first, last = split  # split is not None: the row would have failed above
        passengers.append(
            {
                "_row": excel_row,
                "first_name": first[:100],
                "last_name": last[:100],
                "gender": gender.value if gender else None,
                "dob": dob.isoformat() if dob else None,
                "passenger_type": _passenger_type(dob, travel_date).value,
                "passport_number": passport[:40] or None,
                "passport_expiry": passport_expiry.isoformat() if passport_expiry else None,
                "nationality": _text(cell(row, "Nationality"))[:100] or None,
                "special_services": [],
            }
        )

    if data_rows == 0:
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail="That sheet has a header but no passengers.",
        )

    # _read_journey can fault rows the per-row pass had already cleared — a row
    # can be perfectly well-formed and still contradict the journey every other
    # row states. Counting valid rows inside that loop therefore reported a
    # contradictory sheet as fully valid and let it be submitted, so the tally
    # is taken HERE, once, from the errors that actually exist.
    journey = _read_journey(journey_votes, errors)

    faulted = {e["row"] for e in errors}
    passengers = [p for p in passengers if p["_row"] not in faulted]
    for p in passengers:
        p.pop("_row", None)

    valid_rows = data_rows - len(faulted)
    return {
        "journey": {k: v for k, v in journey.items() if k != "_row"},
        "passengers": passengers,
        "errors": [dict(e) for e in errors],
        "total_rows": data_rows,
        "valid_rows": valid_rows,
        "invalid_rows": len(faulted),
    }


def _passenger_type(dob: datetime.date | None, travel: datetime.date | None) -> PassengerType:
    """Adult / child / infant from the age on the travel date.

    The same thresholds the merchant form's stepper cards state — under 2 is an
    infant, 2-11 a child, 12 and over an adult. Derived rather than asked for,
    because a sheet that carried both a date of birth and a hand-typed type
    would eventually carry two that disagree.
    """
    if dob is None:
        return PassengerType.ADULT
    reference = travel or datetime.date.today()
    years = (
        reference.year
        - dob.year
        - ((reference.month, reference.day) < (dob.month, dob.day))
    )
    if years < 2:
        return PassengerType.INFANT
    if years < 12:
        return PassengerType.CHILD
    return PassengerType.ADULT


def _read_journey(votes: list[dict[str, Any]], errors: list[_Err]) -> dict[str, Any]:
    """The journey the sheet declared, taken from its first valid row.

    Rows are supposed to repeat it identically. Where they do not, the first
    row wins and the disagreement is reported — resolving it silently is how a
    booking ends up flying a sector that appears on none of the rows the
    merchant was looking at.
    """
    if not votes:
        return {}

    first = votes[0]
    compared = ("origin_city", "destination_city", "airline", "flight_number", "travel_date")
    for vote in votes[1:]:
        differing = [k for k in compared if vote.get(k) != first.get(k)]
        if differing:
            errors.append(
                # The vote's OWN row number, not its index in this list: rows
                # that already failed are absent, so an index here would point
                # the merchant at the wrong line of their sheet.
                _Err(
                    vote["_row"],
                    "This row's journey does not match the first row ("
                    + ", ".join(k.replace("_", " ") for k in differing)
                    + "). Every row must carry the same journey.",
                )
            )
    return first


def status_for(valid_rows: int, invalid_rows: int) -> GroupImportStatus:
    if invalid_rows == 0 and valid_rows > 0:
        return GroupImportStatus.VALID
    if valid_rows == 0:
        return GroupImportStatus.INVALID
    return GroupImportStatus.PARTIAL


# ---------------------------------------------------------------------------
# Storing an import
# ---------------------------------------------------------------------------
def _merchant_id_of(actor: User) -> int:
    """The merchant this import belongs to.

    Platform staff have no merchant of their own, and a manifest with no owner
    could not be scoped on read — so this is refused rather than defaulted.
    """
    if not actor.merchant_id:
        raise HTTPException(
            status_code=http_status.HTTP_403_FORBIDDEN,
            detail="Only a merchant can upload a group passenger list.",
        )
    return actor.merchant_id


def _spool(upload: UploadFile) -> tuple[Path, int, str]:
    """Stream the upload to a temp file, capping the size as it goes.

    Never ``await upload.read()`` whole: a 10 MB cap enforced after the bytes
    are already in memory is not a cap. The digest is computed in the same pass
    so the file is walked once.
    """
    digest = hashlib.sha256()
    size = 0
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".upload")
    try:
        while True:
            chunk = upload.file.read(_CHUNK)
            if not chunk:
                break
            size += len(chunk)
            if size > MAX_UPLOAD_BYTES:
                raise HTTPException(
                    status_code=http_status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                    detail=(
                        f"That file is larger than the {MAX_UPLOAD_BYTES // (1024 * 1024)} MB "
                        "limit. Split the party across two uploads."
                    ),
                )
            digest.update(chunk)
            tmp.write(chunk)
        tmp.flush()
    except Exception:
        tmp.close()
        Path(tmp.name).unlink(missing_ok=True)
        raise
    tmp.close()
    if size == 0:
        Path(tmp.name).unlink(missing_ok=True)
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail="That file is empty.",
        )
    return Path(tmp.name), size, digest.hexdigest()


def create_import(
    db: Session,
    actor: User,
    upload: UploadFile,
    journey_type: str,
    *,
    replaces: int | None = None,
) -> GroupBookingImport:
    """Validate an uploaded manifest and record it.

    A sheet with problems is still stored. That is the point of ``partial`` and
    ``invalid``: the merchant needs the error report, and the desk needs to be
    able to see what was actually sent when a merchant says they uploaded
    something. Only :func:`attach_to_request` cares whether it was clean.

    ``replaces`` is "Replace File" — the previous staging row and its bytes go,
    so a merchant cannot accidentally submit against the sheet they just
    replaced.
    """
    _assert_journey_type(journey_type)
    merchant_id = _merchant_id_of(actor)

    if replaces is not None:
        discard(db, actor, replaces)

    tmp_path, size, checksum = _spool(upload)
    try:
        data = tmp_path.read_bytes()
        # Parsing before storing: a file that is not a workbook at all should
        # 400 without leaving bytes behind for nobody.
        result = validate_sheet(parse_rows(data), journey_type)

        key = f"group-imports/{merchant_id}/{uuid.uuid4().hex}.xlsx"
        storage.backend.put(
            key, tmp_path, content_type=_CONTENT_TYPE,
        )
    finally:
        tmp_path.unlink(missing_ok=True)

    imp = GroupBookingImport(
        merchant_id=merchant_id,
        request_id=None,
        journey_type=journey_type,
        original_filename=(upload.filename or "passengers.xlsx")[:255],
        stored_path=key,
        content_type=_CONTENT_TYPE,
        size_bytes=size,
        checksum=checksum,
        imported_passengers=result["passengers"],
        imported_journey=result["journey"],
        validation_errors=result["errors"],
        total_rows=result["total_rows"],
        valid_rows=result["valid_rows"],
        invalid_rows=result["invalid_rows"],
        passengers_imported=len(result["passengers"]),
        validation_status=status_for(result["valid_rows"], result["invalid_rows"]),
        uploaded_by=actor.user_id,
        imported_at=datetime.datetime.now(datetime.timezone.utc),
    )
    db.add(imp)
    db.commit()
    db.refresh(imp)
    return imp


def get(db: Session, actor: User, import_id: int) -> GroupBookingImport:
    """One import, scoped.

    The merchant filter is in the query rather than applied to the result, for
    the reason ``document_service`` states: a scope check after the fetch is one
    forgotten ``if`` away from cross-merchant reads.
    """
    stmt = select(GroupBookingImport).where(GroupBookingImport.import_id == import_id)
    if actor.merchant_id:
        stmt = stmt.where(GroupBookingImport.merchant_id == actor.merchant_id)
    imp = db.execute(stmt).scalar_one_or_none()
    if imp is None:
        raise HTTPException(
            status_code=http_status.HTTP_404_NOT_FOUND,
            detail="That passenger list does not exist.",
        )
    return imp


def for_request(db: Session, actor: User, request_id: int) -> GroupBookingImport:
    """The manifest behind one booking.

    Scoped through ``ticket_service.get_request`` rather than by filtering this
    table directly, so the Admin route obeys exactly the same visibility rule as
    every other request-shaped read — platform staff see every merchant's, a
    merchant sees only its own, and a request that is not visible 404s before
    this table is touched at all.
    """
    from app.services import ticket_service

    request = ticket_service.get_request(db, actor, request_id)
    imp = db.execute(
        select(GroupBookingImport).where(
            GroupBookingImport.request_id == request.request_id
        )
    ).scalar_one_or_none()
    if imp is None:
        raise HTTPException(
            status_code=http_status.HTTP_404_NOT_FOUND,
            detail="This booking was not created from a passenger list.",
        )
    return imp


def discard(db: Session, actor: User, import_id: int) -> None:
    """Drop a staging import and its bytes.

    Refuses once the import belongs to a booking: at that point it is the
    evidence behind a live request, and deleting it would leave the desk
    issuing tickets against a manifest it can no longer open.
    """
    imp = get(db, actor, import_id)
    if imp.request_id is not None:
        raise HTTPException(
            status_code=http_status.HTTP_409_CONFLICT,
            detail="This passenger list belongs to a booking and cannot be replaced.",
        )
    key = imp.stored_path
    db.delete(imp)
    db.commit()
    try:
        storage.backend.delete(key)
    except Exception:
        # The row is gone, which is what callers observe. An orphaned object is
        # a cleanup problem, not a reason to fail the merchant's replace.
        pass


def attach_to_request(db: Session, imp: GroupBookingImport, request: ServiceRequest) -> None:
    """Bind a validated import to the booking raised from it.

    THE GATE. A booking may only be raised from a wholly ``valid`` import — the
    only place that rule is enforced, and the reason ``partial`` exists as a
    distinct state rather than being rounded to one of the other two.
    """
    if imp.validation_status != GroupImportStatus.VALID:
        raise HTTPException(
            status_code=http_status.HTTP_409_CONFLICT,
            detail=(
                f"This passenger list has {imp.invalid_rows} row(s) that still need "
                "fixing. Download the error report, correct them, and upload again."
            ),
        )
    if imp.request_id is not None:
        raise HTTPException(
            status_code=http_status.HTTP_409_CONFLICT,
            detail="This passenger list has already been used for a booking.",
        )
    if imp.merchant_id != request.merchant_id:
        raise HTTPException(
            status_code=http_status.HTTP_403_FORBIDDEN,
            detail="That passenger list belongs to another merchant.",
        )
    imp.request_id = request.request_id
    db.add(imp)


def passengers_of(imp: GroupBookingImport) -> list[dict[str, Any]]:
    """The imported travellers, shaped for ``PassengerInput``."""
    return [dict(p) for p in (imp.imported_passengers or [])]


def open_for_download(imp: GroupBookingImport):
    """The stored bytes, for the merchant or the desk."""
    from app.services.storage import DocumentBytesMissing

    try:
        return storage.backend.open(imp.stored_path)
    except DocumentBytesMissing as exc:
        raise HTTPException(
            status_code=http_status.HTTP_410_GONE,
            detail="The uploaded file is no longer available.",
        ) from exc


# ---------------------------------------------------------------------------
# The error report
# ---------------------------------------------------------------------------
def build_error_report(imp: GroupBookingImport) -> bytes:
    """Every problem, as a sheet the merchant can work through.

    A spreadsheet rather than a text list because the thing being corrected is a
    spreadsheet: the row numbers here line up with the rows in the file still
    open on their screen.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Errors"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 90

    for idx, name in enumerate(("Excel Row", "Column", "What is wrong"), start=1):
        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="8C2F1F")

    for offset, err in enumerate(imp.validation_errors or [], start=2):
        ws.cell(row=offset, column=1, value=err.get("row"))
        ws.cell(row=offset, column=2, value=err.get("column") or "—")
        body = ws.cell(row=offset, column=3, value=err.get("message"))
        body.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"

    summary = wb.create_sheet("Summary")
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 40
    for row, (label, value) in enumerate(
        (
            ("File", imp.original_filename),
            ("Journey type", imp.journey_type.replace("_", " ").title()),
            ("Total rows", imp.total_rows),
            ("Valid rows", imp.valid_rows),
            ("Invalid rows", imp.invalid_rows),
            ("Passengers imported", imp.passengers_imported),
            ("Status", imp.validation_status.value),
        ),
        start=1,
    ):
        summary.cell(row=row, column=1, value=label).font = Font(bold=True)
        summary.cell(row=row, column=2, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def error_report_filename(imp: GroupBookingImport) -> str:
    stem = Path(imp.original_filename).stem[:60] or "passengers"
    return f"{stem}-errors.xlsx"
