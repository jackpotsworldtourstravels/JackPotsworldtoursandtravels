"""CSV/Excel/PDF export — API_CONTRACT.md §6.2.

One row-builder per report type feeds one format registry, so adding a new export format
later is one function plus one registry entry; the router/permission layer above it doesn't
change. CSV needs nothing beyond the standard library; xlsx uses openpyxl; pdf uses reportlab.
"""
import csv
import io
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

Row = dict[str, Any]


def _to_csv(columns: list[tuple[str, str]], rows: list[Row]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([label for _, label in columns])
    for row in rows:
        writer.writerow([row.get(key, "") for key, _ in columns])
    return buf.getvalue().encode("utf-8-sig")  # BOM so Excel opens UTF-8 cleanly


def _to_xlsx(columns: list[tuple[str, str]], rows: list[Row]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append([label for _, label in columns])
    for row in rows:
        ws.append([row.get(key, "") for key, _ in columns])
    for i, (key, label) in enumerate(columns, start=1):
        width = max(len(label), *(len(str(r.get(key, ""))) for r in rows)) if rows else len(label)
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 40)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_pdf(columns: list[tuple[str, str]], rows: list[Row], title: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), title=title)
    styles = getSampleStyleSheet()
    data = [[label for _, label in columns]] + [
        [str(row.get(key, "")) for key, _ in columns] for row in rows
    ]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A2E52")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6FA")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    doc.build([Paragraph(title, styles["Heading2"]), table])
    return buf.getvalue()


FORMATTERS: dict[str, Callable[..., bytes]] = {
    "csv": lambda columns, rows, title: _to_csv(columns, rows),
    "xlsx": lambda columns, rows, title: _to_xlsx(columns, rows),
    "pdf": _to_pdf,
}

MEDIA_TYPES = {
    "csv": "text/csv",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}


def build_export(
    fmt: str, columns: list[tuple[str, str]], rows: list[Row], title: str
) -> tuple[bytes, str]:
    """Returns ``(content_bytes, media_type)``. Raises ``KeyError`` for an unknown format —
    the router validates ``fmt`` against a ``Literal`` first, so this never surfaces to a user."""
    content = FORMATTERS[fmt](columns, rows, title)
    return content, MEDIA_TYPES[fmt]
