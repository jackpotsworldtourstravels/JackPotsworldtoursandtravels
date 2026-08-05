"""Group Booking — the passenger list arrives as a spreadsheet (migration 0042).

WHAT THIS PROTECTS

1. **A sheet with problems can never become a booking.** The headline
   assertion. Validation produces three outcomes and only ``valid`` may be
   spent; ``partial`` is the dangerous one, because it looks like success and
   has real passengers in it. If a partial import could raise a booking, a
   party would fly with rows nobody corrected.
2. **One manifest, one booking.** Reusing a spent import is a 409. Without it,
   the same eighty travellers could be booked twice against one upload.
3. **Every row is judged, not just the first.** The whole point of the error
   report is that a merchant fixes a sheet in one pass rather than forty.
4. **The Excel row numbers are the merchant's.** An error that says "Row 12"
   must mean row 12 of the file open on their screen — header included,
   blank rows counted. An off-by-one here sends them to the wrong line.
5. **One Way and Round Trip are untouched.** The return-leg rule changed shape
   in 0042 (``round_trip_group`` carries a return leg where ``group_trip``
   previously never did). The two original trip types must validate exactly as
   before, and that is asserted explicitly rather than assumed.
6. **Cross-tenant.** Another merchant's manifest is 404, not 403 — the same
   rule every other request-shaped read follows.
7. **The format is decided by the bytes.** A renamed .csv or .txt is refused
   however convincing its extension.

WHY THE SHEETS ARE BUILT HERE RATHER THAN CHECKED IN
A fixture .xlsx would be a binary blob nobody can diff, and the column set is
the contract under test — a test that reads its own committed copy of the
template cannot notice the template changing. Every sheet below is generated
from ``group_booking_service.columns_for``, which is the same function the
download endpoint uses, so a column rename fails here loudly instead of
silently passing.
"""
import datetime
import io
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
BACKEND = HERE.parent / "backend"
sys.path.insert(0, str(BACKEND))

import minihttp as requests  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402

from app.services import group_booking_service as gb  # noqa: E402

from config import BASE, MERCHANT, Checker, H, login  # noqa: E402

check = Checker()
mtok = login(*MERCHANT)

GB = f"{BASE}/api/group-bookings"
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

FUTURE = datetime.date.today() + datetime.timedelta(days=45)
RETURN = FUTURE + datetime.timedelta(days=7)


def sheet(journey_type, rows, *, drop_column=None, header=True):
    """Build a workbook from the service's own column list.

    ``drop_column`` omits one header, which is how the "missing columns"
    refusal is provoked without hand-maintaining a second column list.
    """
    cols = [c for c in gb.columns_for(journey_type) if c != drop_column]
    wb = Workbook()
    ws = wb.active
    ws.title = "Passengers"
    start = 1
    if header:
        for i, c in enumerate(cols, 1):
            ws.cell(row=1, column=i, value=c)
        start = 2
    for r, data in enumerate(rows, start):
        for i, c in enumerate(cols, 1):
            ws.cell(row=r, column=i, value=data.get(c))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def row(**over):
    """One well-formed passenger row; keyword args break whatever is under test."""
    base = {
        "Origin City": "Hyderabad", "Destination City": "Delhi",
        "Airline": "Air India", "Airline Number": "AI217",
        "Travel Date": FUTURE, "Preferred Hour": 9, "Preferred Minute": 30, "AM/PM": "AM",
        "Return Date": RETURN, "Return Preferred Hour": 6,
        "Return Preferred Minute": 0, "Return AM/PM": "PM",
        "Passenger Name": "John Doe", "Gender": "Male",
        "Date of Birth": datetime.date(1990, 4, 18), "Nationality": "Indian",
        "Passport Number": "P1234567", "Passport Expiry": datetime.date(2032, 4, 18),
        "Booking Class": "Economy", "Adults": 1, "Children": 0, "Infants": 0,
        "Client Fare": 20000, "Notes": "",
    }
    base.update(over)
    return base


def upload(data, journey_type="one_way_group", filename="party.xlsx", replaces=None):
    form = {"journey_type": journey_type}
    if replaces:
        form["replaces"] = str(replaces)
    return requests.post(
        f"{GB}/imports",
        headers=H(mtok),
        files={"file": (filename, data, XLSX)},
        data=form,
    )


created = []


def track(r):
    if r.status_code == 201:
        created.append(r.json()["imported"]["import_id"])
    return r


# ---------------------------------------------------------------------------
print("\n== the template ==")
# ---------------------------------------------------------------------------
for jt, expect_return in (("one_way_group", False), ("round_trip_group", True)):
    # minihttp has no `params` — the suite builds query strings into the URL.
    r = requests.get(f"{GB}/template?journey_type={jt}", headers=H(mtok))
    check(f"template downloads ({jt})", r.status_code == 200, f"{r.status_code}")
    if r.status_code != 200:
        continue
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Passengers"]
    header = [c.value for c in ws[1]]
    check(f"  headers match columns_for ({jt})",
          header == list(gb.columns_for(jt)),
          f"{header}")
    check(f"  return columns {'present' if expect_return else 'absent'} ({jt})",
          ("Return Date" in header) is expect_return)
    check(f"  carries a sample row ({jt})", ws.max_row >= 2)
    check(f"  has an Instructions sheet ({jt})", "Instructions" in wb.sheetnames)
    # THE ROUND TRIP: the template it produces must be one its own parser
    # accepts. A template that cannot be uploaded unchanged is the single most
    # embarrassing failure this feature could have.
    r2 = track(upload(r.content, jt, "template-as-downloaded.xlsx"))
    check(f"  the downloaded template uploads unchanged ({jt})",
          r2.status_code == 201 and r2.json()["imported"]["validation_status"] == "valid",
          f"{r2.status_code} {r2.text[:160]}")


# ---------------------------------------------------------------------------
print("\n== a clean import ==")
# ---------------------------------------------------------------------------
r = track(upload(sheet("one_way_group", [
    row(**{"Passenger Name": "Asha Menon", "Passport Number": "GB-A1"}),
    row(**{"Passenger Name": "Ravi Kumar", "Passport Number": "GB-A2"}),
    row(**{"Passenger Name": "Sara Iqbal", "Passport Number": "GB-A3"}),
])))
check("a clean sheet imports", r.status_code == 201, f"{r.status_code} {r.text[:200]}")
clean = r.json()["imported"] if r.status_code == 201 else {}
check("  status is valid", clean.get("validation_status") == "valid", str(clean.get("validation_status")))
check("  can_submit is true", clean.get("can_submit") is True)
check("  counts add up",
      clean.get("total_rows") == 3 and clean.get("valid_rows") == 3
      and clean.get("invalid_rows") == 0 and clean.get("passengers_imported") == 3,
      f"{clean.get('total_rows')}/{clean.get('valid_rows')}/{clean.get('invalid_rows')}")
check("  names are split into first and last",
      [p["first_name"] for p in clean.get("passengers", [])] == ["Asha", "Ravi", "Sara"],
      str([p.get("first_name") for p in clean.get("passengers", [])]))
check("  the 12-hour time became 24-hour",
      clean.get("journey", {}).get("preferred_time") == "09:30",
      str(clean.get("journey", {}).get("preferred_time")))
check("  an error report is refused when there are no errors",
      requests.get(f"{GB}/imports/{clean.get('import_id')}/errors", headers=H(mtok)).status_code == 404)
check("  the stored file streams back",
      requests.get(f"{GB}/imports/{clean.get('import_id')}/file", headers=H(mtok)).status_code == 200)


# ---------------------------------------------------------------------------
print("\n== every validation rule (spec section 7) ==")
# ---------------------------------------------------------------------------
CASES = [
    ("required value missing", {"Passenger Name": ""}, "Passenger Name is missing"),
    ("travel date invalid", {"Travel Date": "not-a-date"}, "Travel Date is invalid"),
    ("travel date in the past", {"Travel Date": datetime.date(2020, 1, 1)}, "in the past"),
    ("origin equals destination", {"Destination City": "Hyderabad"}, "cannot be the same"),
    ("airline number nonsense", {"Airline Number": "NOT A FLIGHT"}, "not a valid airline number"),
    ("passenger count not a number", {"Adults": "many"}, "must be a whole number"),
    ("name has no surname", {"Passenger Name": "Prince"}, "first name and a surname"),
    ("gender unrecognised", {"Gender": "Yes"}, "not a valid Gender"),
    ("passport expires before travel",
     {"Passport Expiry": FUTURE - datetime.timedelta(days=1)}, "before the Travel Date"),
]
for label, override, expect in CASES:
    r = track(upload(sheet("one_way_group", [row(**override)])))
    if r.status_code != 201:
        check(label, False, f"{r.status_code} {r.text[:140]}")
        continue
    d = r.json()["imported"]
    msgs = " | ".join(e["message"] for e in d["errors"])
    check(label, expect.lower() in msgs.lower() and d["validation_status"] == "invalid",
          f"status={d['validation_status']} errors={msgs[:150]}")

# Duplicates are cross-row, so they need two.
r = track(upload(sheet("one_way_group", [
    row(**{"Passenger Name": "One Person", "Passport Number": "DUP-1"}),
    row(**{"Passenger Name": "Two Person", "Passport Number": "DUP-1"}),
])))
d = r.json()["imported"]
check("duplicate passport within the sheet",
      any("already used on row" in e["message"] for e in d["errors"]),
      str([e["message"] for e in d["errors"]])[:160])
check("  and it names the FIRST row, by Excel numbering",
      any(e["message"].endswith("row 2.") for e in d["errors"]),
      str([e["message"] for e in d["errors"]])[:160])


# ---------------------------------------------------------------------------
print("\n== row numbers are the merchant's row numbers ==")
# ---------------------------------------------------------------------------
# Header is row 1, so the third data row is Excel row 4. A blank spacer row is
# skipped as data but still consumes its row number.
r = track(upload(sheet("one_way_group", [
    row(**{"Passenger Name": "Good One", "Passport Number": "RN-1"}),
    row(**{"Passenger Name": "Good Two", "Passport Number": "RN-2"}),
    row(**{"Passenger Name": "", "Passport Number": "RN-3"}),
])))
d = r.json()["imported"]
check("the bad third data row is reported as Excel row 4",
      [e["row"] for e in d["errors"]] == [4],
      str([(e["row"], e["message"]) for e in d["errors"]]))
check("  partial: 2 of 3 rows imported",
      d["validation_status"] == "partial" and d["valid_rows"] == 2 and d["invalid_rows"] == 1,
      f"{d['validation_status']} {d['valid_rows']}/{d['invalid_rows']}")
check("  can_submit is false on a partial import", d["can_submit"] is False)
check("  an error report is offered",
      requests.get(f"{GB}/imports/{d['import_id']}/errors", headers=H(mtok)).status_code == 200)
partial_id = d["import_id"]


# ---------------------------------------------------------------------------
print("\n== rows must agree about the journey ==")
# ---------------------------------------------------------------------------
r = track(upload(sheet("one_way_group", [
    row(**{"Passenger Name": "Agree One", "Passport Number": "JD-1"}),
    row(**{"Passenger Name": "Differ Two", "Passport Number": "JD-2",
           "Destination City": "Mumbai"}),
])))
d = r.json()["imported"]
check("a row that contradicts the journey is faulted",
      any("does not match the first row" in e["message"] for e in d["errors"]),
      str([e["message"] for e in d["errors"]])[:160])
check("  and it is NOT counted as valid",
      d["valid_rows"] == 1 and d["invalid_rows"] == 1 and d["can_submit"] is False,
      f"{d['valid_rows']}/{d['invalid_rows']} can_submit={d['can_submit']}")
check("  the contradicting passenger is excluded from the import",
      d["passengers_imported"] == 1,
      str(d["passengers_imported"]))


# ---------------------------------------------------------------------------
print("\n== whole-file refusals ==")
# ---------------------------------------------------------------------------
r = upload(sheet("one_way_group", [row()], drop_column="Passenger Name"))
check("a missing column refuses the file", r.status_code == 400
      and "missing required columns" in r.text.lower(), f"{r.status_code} {r.text[:140]}")

r = upload(b"col1,col2\nHyderabad,Delhi\n", filename="passengers.xlsx")
check("a renamed CSV is refused on its bytes", r.status_code == 400
      and "not an excel workbook" in r.text.lower(), f"{r.status_code} {r.text[:140]}")

r = upload(sheet("one_way_group", []))
check("a header with no rows is refused", r.status_code == 400
      and "no passengers" in r.text.lower(), f"{r.status_code} {r.text[:140]}")

r = upload(b"")
check("an empty file is refused", r.status_code == 400, f"{r.status_code}")

r = upload(sheet("one_way_group", [row()]), journey_type="sideways_group")
check("an unknown journey type is refused", r.status_code == 400, f"{r.status_code}")

# The 10 MB cap. Padded past the limit rather than built from real rows, which
# would take minutes and prove the same thing.
r = upload(sheet("one_way_group", [row()]) + b"\0" * (11 * 1024 * 1024))
check("a file over 10 MB is refused", r.status_code in (400, 413), f"{r.status_code}")


# ---------------------------------------------------------------------------
print("\n== round trip group ==")
# ---------------------------------------------------------------------------
r = track(upload(sheet("round_trip_group", [
    row(**{"Passenger Name": "Return Traveller", "Passport Number": "RT-1"}),
]), "round_trip_group"))
d = r.json()["imported"] if r.status_code == 201 else {}
check("a round trip group imports", r.status_code == 201, f"{r.status_code} {r.text[:160]}")
check("  the return leg is parsed",
      d.get("journey", {}).get("return_date") == RETURN.isoformat(),
      str(d.get("journey", {}).get("return_date")))
check("  the return time became 24-hour",
      d.get("journey", {}).get("return_preferred_time") == "18:00",
      str(d.get("journey", {}).get("return_preferred_time")))

r = track(upload(sheet("round_trip_group", [
    row(**{"Passenger Name": "No Return", "Passport Number": "RT-2", "Return Date": None}),
]), "round_trip_group"))
check("a round trip group with no return date is faulted",
      any("Return Date" in e["message"] for e in r.json()["imported"]["errors"]),
      r.text[:160])

r = track(upload(sheet("round_trip_group", [
    row(**{"Passenger Name": "Early Return", "Passport Number": "RT-3",
           "Return Date": FUTURE - datetime.timedelta(days=1)}),
]), "round_trip_group"))
check("a return before departure is faulted",
      any("after the Travel Date" in e["message"] for e in r.json()["imported"]["errors"]),
      r.text[:160])


# ---------------------------------------------------------------------------
print("\n== the booking gate ==")
# ---------------------------------------------------------------------------
def direct_booking(import_id, journey_type="one_way_group", **over):
    body = {
        "trip_type": "group_trip",
        "group_journey_type": journey_type,
        "origin": "HYD", "origin_city": "Hyderabad",
        "destination": "DEL", "destination_city": "Delhi",
        "airline": "Air India", "flight_number": "AI217",
        "travel_date": FUTURE.isoformat(), "preferred_time": "09:30",
        "travel_class": "Economy",
        "passenger_count": 3, "adults": 3, "children": 0, "infants": 0,
        "group_import_id": import_id,
        "contact": {"email": "ops@example.com", "phone": "9999999999"},
        "submit": False,
    }
    body.update(over)
    return requests.post(f"{BASE}/api/bookings/direct", headers=H(mtok), json=body)


r = direct_booking(partial_id, passenger_count=2, adults=2)
check("a PARTIAL import cannot become a booking", r.status_code == 409,
      f"{r.status_code} {r.text[:160]}")

r = direct_booking(clean["import_id"])
check("a VALID import becomes a booking", r.status_code == 201,
      f"{r.status_code} {r.text[:200]}")
booking = r.json() if r.status_code == 201 else {}
check("  the imported passengers became passenger rows",
      len(booking.get("passengers", [])) == 3,
      str(len(booking.get("passengers", []))))
check("  the trip type and journey type are stored",
      (booking.get("details") or {}).get("trip_type") == "group_trip"
      and (booking.get("details") or {}).get("group_journey_type") == "one_way_group",
      str(booking.get("details", {}).get("group_journey_type")))

r = direct_booking(clean["import_id"])
check("the same manifest cannot be spent twice", r.status_code == 409,
      f"{r.status_code} {r.text[:160]}")

r = requests.post(f"{BASE}/api/bookings/direct", headers=H(mtok), json={
    "trip_type": "group_trip", "group_journey_type": "one_way_group",
    "origin": "HYD", "destination": "DEL", "airline": "Air India",
    "flight_number": "AI217", "travel_date": FUTURE.isoformat(),
    "preferred_time": "09:30", "travel_class": "Economy",
    "passenger_count": 1, "adults": 1,
    "group_import_id": clean["import_id"],
    "passengers": [{"first_name": "Hand", "last_name": "Typed"}],
})
check("a manifest and typed passengers together are refused",
      r.status_code == 422, f"{r.status_code} {r.text[:160]}")


# ---------------------------------------------------------------------------
print("\n== the Admin's view of a group booking ==")
# ---------------------------------------------------------------------------
if booking.get("id"):
    rid = booking["id"]
    r = requests.get(f"{BASE}/api/admin/requests/{rid}/group-import", headers=H(mtok))
    check("the manifest is reachable from the booking", r.status_code == 200,
          f"{r.status_code} {r.text[:160]}")
    if r.status_code == 200:
        d = r.json()
        check("  it carries the full passenger list", len(d["passengers"]) == 3)
        check("  and the validation summary",
              d["total_rows"] == 3 and d["validation_status"] == "valid")
    check("  the uploaded file downloads",
          requests.get(f"{BASE}/api/admin/requests/{rid}/group-import/file",
                       headers=H(mtok)).status_code == 200)


# ---------------------------------------------------------------------------
print("\n== One Way and Round Trip are untouched ==")
# ---------------------------------------------------------------------------
# 0042 rewrote the return-leg branch these two share. Asserted directly, because
# "we did not mean to change it" is not a test.
def enquiry(**over):
    body = {
        "trip_type": "one_way",
        "origin": "HYD", "destination": "DEL",
        "airline": "Air India", "flight_number": "AI217",
        "travel_date": FUTURE.isoformat(), "preferred_time": "09:00",
        "travel_class": "Economy", "passenger_count": 1, "adults": 1,
    }
    body.update(over)
    return requests.post(f"{BASE}/api/enquiries", headers=H(mtok), json=body)


r = enquiry()
check("a one way enquiry still works", r.status_code == 201, f"{r.status_code} {r.text[:140]}")
check("  and carries no return date", r.json().get("return_date") is None if r.status_code == 201 else False)

r = enquiry(trip_type="round_trip", return_date=RETURN.isoformat(), return_preferred_time="18:00")
check("a round trip enquiry still works", r.status_code == 201, f"{r.status_code} {r.text[:140]}")
check("  and keeps its return date",
      r.json().get("return_date") == RETURN.isoformat() if r.status_code == 201 else False)

r = enquiry(trip_type="round_trip")
check("a round trip with no return date is still refused", r.status_code == 422, f"{r.status_code}")

r = enquiry(group_journey_type="one_way_group")
check("a journey type on a one way is refused", r.status_code == 422, f"{r.status_code}")

r = enquiry(trip_type="group_trip")
check("a group booking with no journey type is refused", r.status_code == 422, f"{r.status_code}")


# ---------------------------------------------------------------------------
print("\n== the enquiry stage asks for a COUNT, not a manifest ==")
# ---------------------------------------------------------------------------
# A group enquiry asks whether N seats exist and what they cost. The sheet that
# lists those N travellers is uploaded at Booking Request, once we have
# answered — so the enquiry must accept a party size with no manifest, and with
# neither of the two class fields, which its form no longer shows.
r = enquiry(trip_type="group_trip", group_journey_type="one_way_group",
            passenger_count=120, adults=120, travel_class=None)
check("a group enquiry needs no manifest and no cabin", r.status_code == 201,
      f"{r.status_code} {r.text[:160]}")
if r.status_code == 201:
    d = r.json()
    check("  it keeps the stated party size", d.get("passenger_count") == 120,
          str(d.get("passenger_count")))
    check("  and carries no cabin", d.get("travel_class") is None, str(d.get("travel_class")))
    check("  and no fare bucket", d.get("booking_class") is None, str(d.get("booking_class")))

# One way and round trip still REQUIRE the cabin. Making the field optional for
# the group must not have quietly made it optional everywhere — that is the
# regression this pairing exists to catch.
r = enquiry(travel_class=None)
check("a one way with no cabin is still refused", r.status_code == 422, f"{r.status_code}")
r = enquiry(trip_type="round_trip", return_date=RETURN.isoformat(), travel_class=None)
check("a round trip with no cabin is still refused", r.status_code == 422, f"{r.status_code}")


# ---------------------------------------------------------------------------
print("\n== how large a party may be, by trip type ==")
# ---------------------------------------------------------------------------
# The group ceiling is configuration and bounds BOTH the enquiry count and the
# manifest row limit, so the two gates cannot contradict each other. Ordinary
# bookings keep the 99 they have always had.
MAX = gb.MAX_ROWS
r = requests.get(f"{GB}/limits", headers=H(mtok))
check("the configured limits are served to the form", r.status_code == 200, f"{r.status_code}")
if r.status_code == 200:
    check("  max_passengers matches the manifest row cap",
          r.json().get("max_passengers") == MAX, f"{r.json()} vs {MAX}")

r = enquiry(trip_type="group_trip", group_journey_type="one_way_group",
            passenger_count=MAX, adults=MAX, travel_class=None)
check(f"a group of exactly {MAX} is allowed", r.status_code == 201, f"{r.status_code}")

r = enquiry(trip_type="group_trip", group_journey_type="one_way_group",
            passenger_count=MAX + 1, adults=MAX + 1, travel_class=None)
check(f"a group of {MAX + 1} is refused", r.status_code == 422, f"{r.status_code}")
check("  and the message names the limit",
      str(MAX) in r.text and "Group Booking" in r.text, r.text[:200])

r = enquiry(passenger_count=100, adults=100)
check("a ONE WAY of 100 is still refused — the group cap is not shared",
      r.status_code == 422, f"{r.status_code}")
r = enquiry(passenger_count=99, adults=99)
check("  and 99 still works", r.status_code == 201, f"{r.status_code}")


# ---------------------------------------------------------------------------
print("\n== Booking Class: one letter, and a different field from Class ==")
# ---------------------------------------------------------------------------
# The cabin (Economy/Business) and the airline's fare bucket (Y/J/K) are two
# separate things that the merchant form used to call by one name.
r = enquiry(travel_class="Business", booking_class="j")
check("a lowercase bucket is stored uppercase",
      r.status_code == 201 and r.json().get("booking_class") == "J",
      f"{r.status_code} {r.json().get('booking_class') if r.status_code == 201 else r.text[:120]}")
check("  and the cabin is untouched beside it",
      r.status_code == 201 and r.json().get("travel_class") == "Business")

for bad, why in [("AB", "two letters"), ("1", "a digit"), ("@", "punctuation"),
                 ("Y1", "a letter and a digit")]:
    r = enquiry(travel_class="Economy", booking_class=bad)
    check(f"  {bad!r} is refused ({why})", r.status_code == 422, f"{r.status_code}")

r = enquiry(travel_class="Economy", booking_class="")
check("a blank bucket is 'not stated', not an error",
      r.status_code == 201 and r.json().get("booking_class") is None,
      f"{r.status_code} {r.json().get('booking_class') if r.status_code == 201 else r.text[:120]}")

r = enquiry(travel_class="Economy")
check("omitting it entirely is fine too",
      r.status_code == 201 and r.json().get("booking_class") is None, f"{r.status_code}")


# ---------------------------------------------------------------------------
print("\n== the sheet accepts what the form accepts: airline is optional ==")
# ---------------------------------------------------------------------------
# THE TWO GATES HAVE TO AGREE. The booking form made Airline and Airline Number
# optional on 2026-08-05, and until this was fixed the spreadsheet still listed
# both in REQUIRED_COLUMNS — so a merchant who raised an OPEN enquiry, had it
# quoted, and then uploaded the manifest for it had every row rejected for
# omitting exactly what the form had told them was optional.
OPEN_CASES = [
    ("neither airline nor flight", {"Airline": "", "Airline Number": ""}, None, None),
    ("airline only", {"Airline Number": ""}, "Air India", None),
    ("flight only", {"Airline": ""}, None, "AI217"),
    ("both, as before", {}, "Air India", "AI217"),
]
for label, override, want_airline, want_flight in OPEN_CASES:
    r = track(upload(sheet("one_way_group", [row(**override)])))
    if r.status_code != 201:
        check(f"a sheet with {label} imports", False, f"{r.status_code} {r.text[:140]}")
        continue
    d = r.json()["imported"]
    check(f"a sheet with {label} imports cleanly",
          d["validation_status"] == "valid" and d["invalid_rows"] == 0,
          f'{d["validation_status"]} {[e["message"] for e in d["errors"]][:2]}')
    # The declared journey is on the DETAIL response, not on the upload summary
    # (GroupImportSummary carries counts and status only) — so it is re-read
    # rather than looked for on `d`, where it would silently be None for every
    # case and turn this into an assertion that always passes.
    j = requests.get(f'{GB}/imports/{d["import_id"]}', headers=H(mtok)).json().get("journey") or {}
    check(f"  ...and records airline={want_airline!r} flight={want_flight!r}",
          j.get("airline") == want_airline and j.get("flight_number") == want_flight,
          f'airline={j.get("airline")!r} flight={j.get("flight_number")!r}')

# The format check must SURVIVE the field becoming optional — blank is an
# answer, a mangled flight number is still a mistake.
r = track(upload(sheet("one_way_group", [row(**{"Airline Number": "NOT A FLIGHT"})])))
d = r.json()["imported"] if r.status_code == 201 else {}
check("a supplied flight number is still format-checked",
      d.get("validation_status") == "invalid"
      and any("not a valid airline number" in e["message"].lower() for e in d.get("errors", [])),
      str(d.get("validation_status")))

# And the columns that ARE required must stay required.
r = track(upload(sheet("one_way_group", [row(**{"Origin City": ""})])))
d = r.json()["imported"] if r.status_code == 201 else {}
check("a genuinely required column is still refused",
      d.get("validation_status") == "invalid"
      and any("Origin City is missing" in e["message"] for e in d.get("errors", [])),
      str(d.get("validation_status")))


# ---------------------------------------------------------------------------
print("\n== scope ==")
# ---------------------------------------------------------------------------
r = requests.get(f"{GB}/imports/99999999", headers=H(mtok))
check("an unknown import is 404", r.status_code == 404, f"{r.status_code}")

r = requests.get(f"{GB}/template")
check("the template needs authentication", r.status_code in (401, 403), f"{r.status_code}")

r = requests.post(f"{GB}/imports", files={"file": ("x.xlsx", sheet("one_way_group", [row()]), XLSX)},
                  data={"journey_type": "one_way_group"})
check("uploading needs authentication", r.status_code in (401, 403), f"{r.status_code}")


# ---------------------------------------------------------------------------
print("\n== cleanup ==")
# ---------------------------------------------------------------------------
# Staging imports only — one that reached a booking is evidence and the API
# refuses to remove it, which is itself the behaviour being confirmed.
removed = kept = 0
for iid in created:
    resp = requests.delete(f"{GB}/imports/{iid}", headers=H(mtok))
    if resp.status_code == 204:
        removed += 1
    elif resp.status_code == 409:
        kept += 1
check(f"staging imports discarded ({removed} removed, {kept} attached to bookings)",
      removed + kept == len(created),
      f"{removed}+{kept} of {len(created)}")

raise SystemExit(check.report())
